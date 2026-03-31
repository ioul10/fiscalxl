"""
core/injector.py  v3 — final
Injection directe des valeurs PDF dans les cellules exactes du template Excel.
Matching par clé normalisée + fuzzy mots-clés + règles contextuelles (mots entiers).
"""

import re, shutil
import openpyxl
from utils.logger import get_logger

logger = get_logger(__name__)

# ─── Cartes des cellules INPUT ────────────────────────────────────────────────
# Bilan Actif  : B=Brut, C=Amort, E=Net N-1
# Bilan Passif : B=Exercice N, C=Exercice N-1
# CPC          : B=Propre N, C=Exercices Précédents, E=Total N-1

ACTIF_CELL_MAP = {
    "frais preliminaires":                          {"B":"B6",  "C":"C6",  "E":"E6"},
    "charges repartir":                             {"B":"B7",  "C":"C7",  "E":"E7"},
    "primes remboursement":                         {"B":"B8",  "C":"C8",  "E":"E8"},
    "recherche developpement":                      {"B":"B10", "C":"C10", "E":"E10"},
    "brevets marques droits":                       {"B":"B11", "C":"C11", "E":"E11"},
    "fonds commercial":                             {"B":"B12", "C":"C12", "E":"E12"},
    "autres immobilisations incorporelles":         {"B":"B13", "C":"C13", "E":"E13"},
    "terrains":                                     {"B":"B15", "C":"C15", "E":"E15"},
    "constructions":                                {"B":"B16", "C":"C16", "E":"E16"},
    "installations techniques":                     {"B":"B17", "C":"C17", "E":"E17"},
    "installations techniques materiel":            {"B":"B17", "C":"C17", "E":"E17"},
    "materiel transport":                           {"B":"B18", "C":"C18", "E":"E18"},
    "materiel de transport":                        {"B":"B18", "C":"C18", "E":"E18"},
    "mobilier mat bureau":                          {"B":"B19", "C":"C19", "E":"E19"},
    "mobilier aménagements":                        {"B":"B19", "C":"C19", "E":"E19"},
    "autres immobilisations corporelles":           {"B":"B20", "C":"C20", "E":"E20"},
    "immobilisations corporelles cours":            {"B":"B21", "C":"C21", "E":"E21"},
    "prets immobilises":                            {"B":"B23", "C":"C23", "E":"E23"},
    "autres creances financieres":                  {"B":"B24", "C":"C24", "E":"E24"},
    "titres participation":                         {"B":"B25", "C":"C25", "E":"E25"},
    "autres titres immobilises":                    {"B":"B26", "C":"C26", "E":"E26"},
    "ecarts conversion actif e":                    {"B":"B27", "C":"C27", "E":"E27"},
    "diminution creances immobilisees":             {"B":"B28", "C":"C28", "E":"E28"},
    "augmentation dettes financieres":              {"B":"B29", "C":"C29", "E":"E29"},
    "marchandises":                                 {"B":"B33", "C":"C33", "E":"E33"},
    "matieres fournitures consommables":            {"B":"B34", "C":"C34", "E":"E34"},
    "produits cours":                               {"B":"B35", "C":"C35", "E":"E35"},
    "produits intermediaires residuels":            {"B":"B36", "C":"C36", "E":"E36"},
    "produits finis":                               {"B":"B37", "C":"C37", "E":"E37"},
    "fournisseurs debiteurs avances":               {"B":"B39", "C":"C39", "E":"E39"},
    "clients comptes rattaches":                    {"B":"B40", "C":"C40", "E":"E40"},
    "personnel actif circulant":                    {"B":"B41", "C":"C41", "E":"E41"},
    "etat actif circulant":                         {"B":"B42", "C":"C42", "E":"E42"},
    "comptes associes actif":                       {"B":"B43", "C":"C43", "E":"E43"},
    "autres debiteurs":                             {"B":"B44", "C":"C44", "E":"E44"},
    "comptes regularisation actif":                 {"B":"B45", "C":"C45", "E":"E45"},
    "titres valeurs placement":                     {"B":"B46", "C":"C46", "E":"E46"},
    "ecarts conversion circulants":                 {"B":"B47", "C":"C47", "E":"E47"},
    "cheques valeurs encaisser":                    {"B":"B50", "C":"C50", "E":"E50"},
    "banques tg ccp":                               {"B":"B51", "C":"C51", "E":"E51"},
    "caisse regie avances":                         {"B":"B52", "C":"C52", "E":"E52"},
}

PASSIF_CELL_MAP = {
    "capital social":                               {"B":"B6",  "C":"C6"},
    "moins actionnaires":                           {"B":"B7",  "C":"C7"},
    "capital appele":                               {"B":"B8",  "C":"C8"},
    "prime emission fusion":                        {"B":"B9",  "C":"C9"},
    "ecarts reevaluation":                          {"B":"B10", "C":"C10"},
    "reserve legale":                               {"B":"B11", "C":"C11"},
    "autres reserves":                              {"B":"B12", "C":"C12"},
    "report nouveau":                               {"B":"B13", "C":"C13"},
    "resultat instance affectation":                {"B":"B14", "C":"C14"},
    "resultat net exercice passif":                 {"B":"B15", "C":"C15"},
    "subventions investissement passif":            {"B":"B18", "C":"C18"},
    "provisions reglementees":                      {"B":"B19", "C":"C19"},
    "dettes financement":                           {"B":"B20", "C":"C20"},
    "emprunts obligataires":                        {"B":"B21", "C":"C21"},
    "autres dettes financement":                    {"B":"B22", "C":"C22"},
    "provisions durables risques":                  {"B":"B23", "C":"C23"},
    "provisions risques":                           {"B":"B24", "C":"C24"},
    "provisions charges":                           {"B":"B25", "C":"C25"},
    "ecarts conversion passif e":                   {"B":"B26", "C":"C26"},
    "augmentation creances immobilisees":           {"B":"B27", "C":"C27"},
    "diminution dettes financement":                {"B":"B28", "C":"C28"},
    "fournisseurs comptes rattaches passif":        {"B":"B32", "C":"C32"},
    "clients crediteurs avances passif":            {"B":"B33", "C":"C33"},
    "personnel passif":                             {"B":"B34", "C":"C34"},
    "organismes sociaux":                           {"B":"B35", "C":"C35"},
    "etat passif":                                  {"B":"B36", "C":"C36"},
    "comptes associes passif":                      {"B":"B37", "C":"C37"},
    "autres creanciers":                            {"B":"B38", "C":"C38"},
    "comptes regularisation passif":                {"B":"B39", "C":"C39"},
    "autres provisions risques charges":            {"B":"B40", "C":"C40"},
    "ecarts conversion passif circulants":          {"B":"B41", "C":"C41"},
    "credits escompte":                             {"B":"B44", "C":"C44"},
    "credits tresorerie":                           {"B":"B45", "C":"C45"},
    "banques soldes crediteurs":                    {"B":"B46", "C":"C46"},
}

CPC_CELL_MAP = {
    "ventes marchandises etat":                     {"B":"B5",  "C":"C5",  "E":"E5"},
    "ventes biens services":                        {"B":"B6",  "C":"C6",  "E":"E6"},
    "chiffres affaires":                            {"B":"B6",  "C":"C6",  "E":"E6"},
    "chiffre affaires":                             {"B":"B6",  "C":"C6",  "E":"E6"},
    "variation stocks produits":                    {"B":"B8",  "C":"C8",  "E":"E8"},
    "immobilisations produites entreprise":         {"B":"B9",  "C":"C9",  "E":"E9"},
    "subventions exploitation":                     {"B":"B10", "C":"C10", "E":"E10"},
    "autres produits exploitation":                 {"B":"B11", "C":"C11", "E":"E11"},
    "reprises exploitation transferts":             {"B":"B12", "C":"C12", "E":"E12"},
    "achats revendus marchandises":                 {"B":"B15", "C":"C15", "E":"E15"},
    "achats consommes matieres":                    {"B":"B16", "C":"C16", "E":"E16"},
    "autres charges externes":                      {"B":"B17", "C":"C17", "E":"E17"},
    "impots taxes":                                 {"B":"B18", "C":"C18", "E":"E18"},
    "charges personnel":                            {"B":"B19", "C":"C19", "E":"E19"},
    "autres charges exploitation":                  {"B":"B20", "C":"C20", "E":"E20"},
    "dotations exploitation":                       {"B":"B21", "C":"C21", "E":"E21"},
    "produits titres participation":                {"B":"B25", "C":"C25", "E":"E25"},
    "gains change":                                 {"B":"B26", "C":"C26", "E":"E26"},
    "interets autres produits financiers":          {"B":"B27", "C":"C27", "E":"E27"},
    "reprises financieres transferts":              {"B":"B28", "C":"C28", "E":"E28"},
    "charges interets":                             {"B":"B31", "C":"C31", "E":"E31"},
    "pertes change":                                {"B":"B32", "C":"C32", "E":"E32"},
    "autres charges financieres":                   {"B":"B33", "C":"C33", "E":"E33"},
    "dotations financieres":                        {"B":"B34", "C":"C34", "E":"E34"},
    "produits cessions immobilisations":            {"B":"B39", "C":"C39", "E":"E39"},
    "subventions equilibre":                        {"B":"B40", "C":"C40", "E":"E40"},
    "reprises subventions investissement":          {"B":"B41", "C":"C41", "E":"E41"},
    "autres produits courants":                     {"B":"B42", "C":"C42", "E":"E42"},
    "reprises courantes transferts":                {"B":"B43", "C":"C43", "E":"E43"},
    "valeurs nettes amort immobilisations":         {"B":"B46", "C":"C46", "E":"E46"},
    "subventions accordees":                        {"B":"B47", "C":"C47", "E":"E47"},
    "autres charges courantes":                     {"B":"B48", "C":"C48", "E":"E48"},
    "dotations courantes amort provisions":         {"B":"B49", "C":"C49", "E":"E49"},
    "impots benefices":                             {"B":"B53", "C":"C53", "E":"E53"},
}

# ─── Labels de section/total → NE PAS injecter ───────────────────────────────
# Testés via startswith(norm) pour attraper toutes les variantes

NO_INJECT_STARTSWITH = (
    # CPC : résultats calculés par formules
    "resultat d exploitation",
    "resultat financier",
    "resultat courant",
    "resultat non courant",
    "resultat avant impots",
    "resultat net xi",
    "resultat net total",
    # Totaux explicites
    "total i ",  "total ii", "total iii", "total des ", "total general",
    "total vii", "total viii",
    # Sections sans cellules numériques
    "charges non courant",
    "produits non courant",
    "produits d exploitation",
    "charges d exploitation",
    "charges financier",
    "produits financier",
    "capitaux propres assimil",
    "dettes du passif circulant",
)

NO_INJECT_EXACT = {
    "iii", "vii",
    "dont verse",
    "capital appele",
    "immobilisations financieres d",
    "stocks f",
    "creances actif circulant g",
    "total a b c d e",
    "tresorerie actif",
    "dettes passif circulant",
    "total general i ii iii",
    # Totaux CPC intermédiaires (calculés par formule)
    "total iv", "total v", "total viii", "total ix",
}

# ─── Désambiguïsation contextuelle par mots entiers ──────────────────────────
# Règle : le label normalisé doit CONTENIR le mot-clé comme mot entier
# (pas comme sous-chaîne d'un autre mot)

CONTEXT_WORD_RULES = {
    # (mot_clé, section) → clé_dans_cell_map
    # Un label "Personnel" seul → actif ou passif selon section
    # Un label "Capital social ou Personnel" → capital social (pas personnel)
    ("personnel", "actif"):  ("personnel actif circulant",  None),
    ("personnel", "passif"): ("personnel passif",           None),
    ("etat",      "actif"):  ("etat actif circulant",       None),
    ("etat",      "passif"): ("etat passif",                None),
    ("fournisseurs comptes rattaches", "passif"): ("fournisseurs comptes rattaches passif", None),
    ("comptes regularisation",         "passif"): ("comptes regularisation passif",         None),
    ("comptes associes",               "passif"): ("comptes associes passif",               None),
    ("clients crediteurs",             "passif"): ("clients crediteurs avances passif",     None),
}

# Labels passif qui ont des alias spéciaux
PASSIF_LABEL_MAP = {
    # label_normalisé_pdf → clé_dans_PASSIF_CELL_MAP
    "capital social ou personnel 1":    "capital social",
    "capital social ou personnel":      "capital social",
    "resultat net de l exercice":       "resultat net exercice passif",
    "resultat net de l exercice 2":     "resultat net exercice passif",
    "subvention d investissement":      "subventions investissement passif",
    "subventions d investissement":     "subventions investissement passif",
    "resultat en instance d affectation": "resultat instance affectation",
    "report a nouveau 2":               "report nouveau",
    "report a nouveau":                 "report nouveau",
    "fournisseurs et comptes rattaches": "fournisseurs comptes rattaches passif",
    "comptes de regularisation passif": "comptes regularisation passif",
    "autres provisions pour risques et charges g": "autres provisions risques charges",
}

CPC_LABEL_MAP = {
    "interets et autres produits fi":       "interets autres produits financiers",
    "interets et autres produits financiers":"interets autres produits financiers",
    "achats consommes 2 de matieres et fournitures": "achats consommes matieres",
    "valeurs nettes d amortissements des immobilisations cedees": "valeurs nettes amort immobilisations",
    "dotations non courantes aux amortissements et aux provisions": "dotations courantes amort provisions",
    "reprises d exploitation transferts de charges": "reprises exploitation transferts",
    "reprises non courantes transferts de charges": "reprises courantes transferts",
    "reprises financieres transferts de charges": "reprises financieres transferts",
    "achats revendus 2 de marchandises": "achats revendus marchandises",
    "chiffres d affaires": "chiffres affaires",
    "autres produits non courants": "autres produits courants",
    "autres charges non courantes": "autres charges courantes",
}


class TemplateInjector:

    def __init__(self, template_path: str):
        self.template_path = template_path

    def inject(self, extracted: dict, output_path: str) -> dict:
        shutil.copy2(self.template_path, output_path)
        wb = openpyxl.load_workbook(output_path)
        stats = {"injected": 0, "not_found": []}

        self._inject_info(wb, extracted.get("info", {}))

        for section, ws_name, cell_map, col_order, label_map in [
            ("actif",  "2 - Bilan Actif",  ACTIF_CELL_MAP,  ["B","C","E"], {}),
            ("passif", "3 - Bilan Passif", PASSIF_CELL_MAP, ["B","C"],     PASSIF_LABEL_MAP),
            ("cpc",    "4 - CPC",          CPC_CELL_MAP,    ["B","C","E"], CPC_LABEL_MAP),
        ]:
            key_name = f"{section}_values"
            n = self._inject_section(
                wb[ws_name],
                extracted.get(key_name, {}),
                cell_map,
                section=section,
                col_order=col_order,
                label_map=label_map,
            )
            stats["injected"] += n["injected"]
            stats["not_found"].extend(n["not_found"])

        wb.save(output_path)
        logger.info(f"Injection : {stats['injected']} valeurs · {len(stats['not_found'])} non mappés")
        if stats["not_found"]:
            logger.warning(f"Non mappés : {stats['not_found'][:8]}")
        return stats

    def _inject_info(self, wb, info: dict):
        ws = wb["1 - Infos Générales"]
        for row, key in {4:"raison_sociale", 5:"taxe_professionnelle",
                         6:"identifiant_fiscal", 7:"adresse",
                         8:"exercice", 9:"date_declaration"}.items():
            v = info.get(key, "")
            if v:
                ws.cell(row=row, column=2).value = str(v)

    def _inject_section(self, ws, values: dict, cell_map: dict,
                        section: str, col_order: list,
                        label_map: dict) -> dict:
        injected = 0
        not_found = []
        idx = build_index(cell_map)
        passif_idx = build_index(PASSIF_CELL_MAP)

        for raw_label, vals in values.items():
            norm = normalize(raw_label)

            # ── Filtres NO_INJECT ──
            if norm in NO_INJECT_EXACT:
                continue
            if any(norm.startswith(p) for p in NO_INJECT_STARTSWITH):
                continue

            # ── Filtre anti-contamination actif ──
            # Dans la section actif : si Brut=None, toujours skip
            # (Net=Brut-Amort par formule → sans Brut, pas d'injection utile)
            if section == "actif":
                brut = vals[0] if vals else None
                if brut is None:
                    continue  # skip silencieux, pas de not_found

            # ── Résolution du label ──
            key = self._resolve_label(norm, section, idx, cell_map, label_map)

            if not key:
                not_found.append(raw_label[:45])
                continue

            # ── Injection ──
            refs = cell_map[key]
            for i, col in enumerate(col_order):
                if col in refs and i < len(vals) and vals[i] is not None:
                    r = int(refs[col][1:])
                    c = ord(refs[col][0]) - 64
                    ws.cell(row=r, column=c).value = vals[i]
                    injected += 1

        return {"injected": injected, "not_found": not_found}

    def _resolve_label(self, norm: str, section: str,
                       idx: dict, cell_map: dict, label_map: dict) -> str | None:
        """
        Résout un label normalisé en clé du cell_map.
        Ordre de priorité :
          1. label_map direct (alias explicites)
          2. Règles contextuelles par mot entier
          3. Correspondance exacte dans idx
          4. Fuzzy match (seuil 0.45)
        """
        # 1. Alias explicite
        if norm in label_map:
            return label_map[norm]

        # 2. Règles contextuelles (mot entier uniquement)
        norm_words = set(re.findall(r'\b\w+\b', norm))
        for (kw, sec), (mapped_key, _) in CONTEXT_WORD_RULES.items():
            if sec == section:
                kw_words = set(re.findall(r'\b\w+\b', kw))
                # Tous les mots du mot-clé doivent être dans norm
                # ET norm ne doit pas contenir d'autres qualificatifs qui changent le sens
                if kw_words.issubset(norm_words):
                    # Vérifier que le label n'est pas plus spécifique (ex: "capital social ou personnel")
                    if mapped_key in cell_map:
                        # Exclure si label contient des mots indicatifs d'un autre poste
                        exclude_if = {
                            ("personnel", "passif"): {"capital", "social"},
                            ("etat",      "passif"): {"capital", "social"},
                            ("etat",      "actif"):  {"capital", "social"},
                        }
                        excl = exclude_if.get((kw, sec), set())
                        if not excl.intersection(norm_words):
                            return mapped_key

        # 3. Exact
        if norm in idx:
            return idx[norm]

        # 4. Fuzzy
        return find_key_fuzzy(norm, idx, cell_map)


# ─── Utilitaires ──────────────────────────────────────────────────────────────

def normalize(s: str) -> str:
    s = str(s).lower().strip()
    for k, v in {"é":"e","è":"e","ê":"e","à":"a","â":"a","ô":"o","û":"u",
                 "î":"i","ç":"c","œ":"oe","ë":"e","ï":"i","ù":"u","ü":"u"}.items():
        s = s.replace(k, v)
    s = re.sub(r"[^\w\s]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def keywords(s: str) -> set:
    stop = {"les","des","de","du","et","en","sur","par","pour","aux","une","un",
            "la","le","au","ou","est","sont","ne","se","dont","avec","sans",
            "dans","sous","vers","non","plus","a","b","c","d","e","f","g","h",
            "i","j","par","au","aux","en","sur","pour","et","ou"}
    return {w for w in re.findall(r"[a-z]{3,}", s) if w not in stop}


def build_index(cell_map: dict) -> dict:
    return {normalize(k): k for k in cell_map}


def find_key(norm: str, index: dict, cell_map: dict) -> str | None:
    if norm in index:
        return index[norm]
    return find_key_fuzzy(norm, index, cell_map)


def find_key_fuzzy(norm: str, index: dict, cell_map: dict,
                   threshold: float = 0.45) -> str | None:
    norm_kw = keywords(norm)
    if not norm_kw:
        return None
    best, best_score = None, 0
    for k_norm, k_orig in index.items():
        k_kw = keywords(k_norm)
        if not k_kw:
            continue
        common = norm_kw & k_kw
        score = len(common) / max(len(norm_kw), len(k_kw))
        if score > best_score and score >= threshold:
            best_score = score
            best = k_orig
    return best
