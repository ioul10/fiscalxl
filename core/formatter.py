"""
core/formatter.py
Génère l'Excel final unifié depuis la structure normalisée.
Même rendu peu importe le format source (AMMC ou DGI).
Style : bordures fines, gras pour sections/totaux/résultats,
indentation pour sous-postes. Pas de couleurs de fond.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

NUM_FMT  = '#,##0.00;(#,##0.00);"-"'
FONT     = 'Arial'

# Styles
S_THIN   = Side(style='thin',   color='000000')
S_THICK  = Side(style='medium', color='000000')

def _border(top='thin', bottom='thin', left='thin', right='thin'):
    def s(t): return Side(style=t, color='000000') if t else Side(style=None)
    return Border(top=s(top), bottom=s(bottom), left=s(left), right=s(right))

def _cell(ws, r, c, value=None, bold=False, size=9, align='left',
          indent=0, num_fmt=None, border='thin', wrap=True, italic=False):
    cell = ws.cell(row=r, column=c, value=value)
    cell.font      = Font(name=FONT, size=size, bold=bold, italic=italic)
    cell.alignment = Alignment(horizontal=align, vertical='center',
                               indent=indent, wrap_text=wrap)
    cell.border    = _border() if border == 'thin' else _border(
        top='medium', bottom='medium', left='medium', right='medium')
    if num_fmt and value is not None:
        cell.number_format = num_fmt
    return cell

def _row_styles(typ: str) -> dict:
    """Retourne les paramètres de style selon le type de ligne."""
    if typ == 'total':
        return {'bold': True, 'border': 'thick', 'indent': 0}
    if typ == 'result':
        return {'bold': True, 'border': 'thin', 'indent': 0}
    if typ == 'section':
        return {'bold': True, 'border': 'thin', 'indent': 0}
    if typ == 'subtotal':
        return {'bold': True, 'border': 'thin', 'indent': 0}
    return {'bold': False, 'border': 'thin', 'indent': 1}


# ── Feuille Identification ────────────────────────────────────────────────────

def _sheet_ident(wb, info: dict):
    ws = wb.create_sheet('1 — Identification')
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 55

    ws.merge_cells('A1:B1')
    c = ws.cell(1, 1, 'PIÈCES ANNEXES À LA DÉCLARATION FISCALE')
    c.font = Font(name=FONT, bold=True, size=13)
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = _border(top='medium', bottom='medium',
                       left='medium', right='medium')
    ws.row_dimensions[1].height = 26

    ws.merge_cells('A2:B2')
    c = ws.cell(2, 1, "IMPÔTS SUR LES SOCIÉTÉS — Modèle Comptable Normal (loi 9-88)")
    c.font = Font(name=FONT, italic=True, size=10)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 18

    # En-têtes tableau
    for col, txt in [(1, 'Champ'), (2, 'Valeur')]:
        c = ws.cell(4, col, txt)
        c.font   = Font(name=FONT, bold=True, size=9)
        c.border = _border(top='medium', bottom='medium',
                           left='medium', right='medium')
        c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[4].height = 16

    fields = [
        ('Raison sociale',       info.get('raison_sociale', '')),
        ('Identifiant fiscal',   info.get('identifiant_fiscal', '')),
        ('Taxe professionnelle', info.get('taxe_professionnelle', '')),
        ('Adresse',              info.get('adresse', '')),
        ('Exercice',             info.get('exercice', '')),
    ]
    for i, (lbl, val) in enumerate(fields, 5):
        ws.row_dimensions[i].height = 16
        c1 = ws.cell(i, 1, lbl)
        c1.font = Font(name=FONT, bold=True, size=9)
        c1.border = _border()
        c1.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        c2 = ws.cell(i, 2, val)
        c2.font = Font(name=FONT, size=9)
        c2.border = _border()
        c2.alignment = Alignment(horizontal='left', vertical='center', indent=1)


# ── Feuille Bilan Actif ───────────────────────────────────────────────────────

def _sheet_actif(wb, info: dict, rows: list):
    ws = wb.create_sheet('2 — Bilan Actif')
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 48
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18

    # Titre
    ws.merge_cells('A1:E1')
    c = ws.cell(1, 1, f"BILAN ACTIF — {info.get('raison_sociale','')} — {info.get('exercice','')}")
    c.font = Font(name=FONT, bold=True, size=11)
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = _border(top='medium', bottom='medium', left='medium', right='medium')
    ws.row_dimensions[1].height = 22

    # En-têtes colonnes
    headers = ['DÉSIGNATION', 'BRUT', 'AMORT. & PROV.', 'NET EXERCICE N', 'NET EXERCICE N-1']
    for col, h in enumerate(headers, 1):
        c = ws.cell(2, col, h)
        c.font      = Font(name=FONT, bold=True, size=9)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border    = _border(top='medium', bottom='medium', left='medium', right='medium')
    ws.row_dimensions[2].height = 28
    ws.freeze_panes = 'A3'

    r = 3
    for d in rows:
        st = _row_styles(d.get('type', 'normal'))
        ws.row_dimensions[r].height = 14 if d['type'] == 'normal' else 16
        _cell(ws, r, 1, d['label'],      bold=st['bold'], indent=st['indent'],
              border=st['border'])
        _cell(ws, r, 2, d.get('brut'),   bold=st['bold'], align='right',
              num_fmt=NUM_FMT, border=st['border'])
        _cell(ws, r, 3, d.get('amort'),  bold=st['bold'], align='right',
              num_fmt=NUM_FMT, border=st['border'])
        _cell(ws, r, 4, d.get('net_n'),  bold=st['bold'], align='right',
              num_fmt=NUM_FMT, border=st['border'])
        _cell(ws, r, 5, d.get('net_n1'), bold=st['bold'], align='right',
              num_fmt=NUM_FMT, border=st['border'])
        r += 1


# ── Feuille Bilan Passif ──────────────────────────────────────────────────────

def _sheet_passif(wb, info: dict, rows: list):
    ws = wb.create_sheet('3 — Bilan Passif')
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 52
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20

    ws.merge_cells('A1:C1')
    c = ws.cell(1, 1, f"BILAN PASSIF — {info.get('raison_sociale','')} — {info.get('exercice','')}")
    c.font = Font(name=FONT, bold=True, size=11)
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = _border(top='medium', bottom='medium', left='medium', right='medium')
    ws.row_dimensions[1].height = 22

    for col, h in enumerate(['DÉSIGNATION', 'EXERCICE N', 'EXERCICE N-1'], 1):
        c = ws.cell(2, col, h)
        c.font      = Font(name=FONT, bold=True, size=9)
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border    = _border(top='medium', bottom='medium', left='medium', right='medium')
    ws.row_dimensions[2].height = 20
    ws.freeze_panes = 'A3'

    r = 3
    for d in rows:
        st = _row_styles(d.get('type', 'normal'))
        ws.row_dimensions[r].height = 14 if d['type'] == 'normal' else 16
        _cell(ws, r, 1, d['label'],       bold=st['bold'], indent=st['indent'],
              border=st['border'])
        _cell(ws, r, 2, d.get('val_n'),   bold=st['bold'], align='right',
              num_fmt=NUM_FMT, border=st['border'])
        _cell(ws, r, 3, d.get('val_n1'),  bold=st['bold'], align='right',
              num_fmt=NUM_FMT, border=st['border'])
        r += 1

    # Note bas de page
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    c = ws.cell(r, 1, '(1) Capital personnel débiteur.  (2) Bénéficiaire (+) / Déficitaire (−).')
    c.font = Font(name=FONT, italic=True, size=8)
    c.alignment = Alignment(horizontal='left', vertical='center', indent=1)


# ── Feuille CPC ───────────────────────────────────────────────────────────────

def _sheet_cpc(wb, info: dict, rows: list):
    ws = wb.create_sheet('4 — CPC')
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 52
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18

    ws.merge_cells('A1:E1')
    c = ws.cell(1, 1, f"COMPTE DE PRODUITS ET CHARGES — {info.get('raison_sociale','')} — {info.get('exercice','')}")
    c.font = Font(name=FONT, bold=True, size=11)
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = _border(top='medium', bottom='medium', left='medium', right='medium')
    ws.row_dimensions[1].height = 22

    headers = ['DÉSIGNATION', "PROPRES À L'EXERCICE",
               'EXERCICES PRÉCÉDENTS', 'TOTAUX EXERCICE N', 'TOTAUX EXERCICE N-1']
    for col, h in enumerate(headers, 1):
        c = ws.cell(2, col, h)
        c.font      = Font(name=FONT, bold=True, size=9)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border    = _border(top='medium', bottom='medium', left='medium', right='medium')
    ws.row_dimensions[2].height = 30
    ws.freeze_panes = 'A3'

    r = 3
    for d in rows:
        st = _row_styles(d.get('type', 'normal'))
        ws.row_dimensions[r].height = 14 if d['type'] == 'normal' else 16
        _cell(ws, r, 1, d['label'],          bold=st['bold'], indent=st['indent'],
              border=st['border'])
        _cell(ws, r, 2, d.get('propre_n'),   bold=st['bold'], align='right',
              num_fmt=NUM_FMT, border=st['border'])
        _cell(ws, r, 3, d.get('prec_n'),     bold=st['bold'], align='right',
              num_fmt=NUM_FMT, border=st['border'])
        _cell(ws, r, 4, d.get('total_n'),    bold=st['bold'], align='right',
              num_fmt=NUM_FMT, border=st['border'])
        _cell(ws, r, 5, d.get('total_n1'),   bold=st['bold'], align='right',
              num_fmt=NUM_FMT, border=st['border'])
        r += 1

    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    c = ws.cell(r, 1, '(1) Stock final − Stock initial.  (2) Achats − Variation de stock.')
    c.font = Font(name=FONT, italic=True, size=8)
    c.alignment = Alignment(horizontal='left', vertical='center', indent=1)


# ── Point d'entrée ────────────────────────────────────────────────────────────

def build_excel(data: dict, output_path: str):
    """
    Construit l'Excel final depuis la structure normalisée.
    data = {info, actif, passif, cpc}
    """
    wb = Workbook()
    wb.remove(wb.active)

    _sheet_ident(wb,  data['info'])
    _sheet_actif(wb,  data['info'], data['actif'])
    _sheet_passif(wb, data['info'], data['passif'])
    _sheet_cpc(wb,    data['info'], data['cpc'])

    wb.save(output_path)
