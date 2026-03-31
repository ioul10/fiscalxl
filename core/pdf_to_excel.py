"""
core/pdf_to_excel.py  v2
Extraction PDF fiscal → Excel structuré (MCN loi 9-88 Maroc).

Approche : extract_words() avec coordonnées bbox.
Chaque mot a une position (x0, top). On groupe par ligne (top),
puis on sépare label (x faible) et valeurs (x fort) selon des
seuils calibrés sur la largeur de page (en %).
Compatible avec tous les PDFs MCN, indépendamment de leur structure interne.
"""

import re
import unicodedata
from collections import defaultdict

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ── Palette ──────────────────────────────────────────────────────────────────
C_DARK_BLUE  = "1F3864"
C_MED_BLUE   = "2E75B6"
C_LIGHT_BLUE = "BDD7EE"
C_SECTION    = "D6E4F0"
C_SUBTOTAL   = "EBF3FB"
C_RESULT     = "2E4057"
C_WHITE      = "FFFFFF"
C_GRAY_BG    = "F5F7FA"
C_BORDER     = "B8CCE4"

NUM_FMT = '#,##0.00;(#,##0.00);"-"'

# ── Mots-clés détection sections ─────────────────────────────────────────────
ACTIF_KW  = ["actif immobilise", "immobilisations en non valeur",
              "bilan (actif", "bilan actif", "immobilisation incorporelle",
              "frais preliminaires", "b i l a n (actif"]
PASSIF_KW = ["capitaux propres", "bilan (passif", "bilan passif",
              "capital social", "dettes de financement", "b i l a n (passif"]
CPC_KW    = ["produits exploitation", "charges exploitation",
             "compte de produits", "ventes de marchandises",
             "chiffre d affaires", "produits d exploitation"]

# ── Labels à ignorer ─────────────────────────────────────────────────────────
SKIP_EXACT = {
    "brut", "net", "designation", "operations",
    "a c t i f", "p a s s i f", "b i l a n",
    "propres a l exercice", "concernant les exercices precedents",
    "totaux de l exercice", "totaux de l exercice precedent",
    "amortissements et provisions", "exercice precedent",
    "1", "2", "3 = 2 + 1", "4", "3 = 1 + 2",
}
SKIP_PREFIX = (
    "tableau n", "bilan (", "compte de produits", "societe ",
    "identifiant", "exercice du", "cadre reserve", "signature",
    "nb :", "1)variation", "2)achat", "modele normal",
    "fes le", "casablanca",
)
SKIP_SUFFIX = ("(1/2)", "(2/2)", "(hors taxes)", "(suite)")

# ── Types de ligne ────────────────────────────────────────────────────────────
TOTAL_KW = ("total i ", "total ii", "total iii", "total general",
            "total (a+b", "total i+ii", "total iv", "total v ",
            "total vi", "total vii", "total viii", "total ix",
            "total xiv", "total xv", "total des capitaux",
            "total des produits", "total des charges")
RESULT_KW = ("resultat d exploitation", "resultat financier",
             "resultat courant", "resultat non courant",
             "resultat avant impot", "resultat net",
             "impots sur les", "impot sur les")
SECTION_KW = ("produits d exploitation", "charges d exploitation",
              "produits financiers", "charges financieres",
              "produits non courant", "charges non courant",
              "capitaux propres assimile", "dettes de financement",
              "provisions durables", "dettes du passif circulant",
              "tresorerie", "ecarts de conversion",
              "immobilisations en non", "immobilisation incorporelle",
              "immobilisations corporelles", "immobilisations financiere",
              "stocks", "creances de l actif", "titres et valeurs")


# ═══════════════════════════════════════════════════════════════════════════════
# UTILITAIRES
# ═══════════════════════════════════════════════════════════════════════════════

def _norm(s: str) -> str:
    s = unicodedata.normalize('NFD', s)
    s = s.encode('ascii', 'ignore').decode('utf-8').lower()
    s = re.sub(r"[^\w\s]", " ", s)
    return re.sub(r"\s+", " ", s).strip()

def _parse_num_tokens(tokens: list):
    if not tokens:
        return None
    s = "".join(tokens).replace(" ", "").replace("\xa0", "")
    if not s or s in ["-", "—"]:
        return None
    neg = s.startswith("(") and s.endswith(")")
    if neg:
        s = s[1:-1]
    if s.startswith("-"):
        neg, s = True, s[1:]
    s = s.replace(",", ".")
    try:
        v = float(s)
        return -v if neg else v
    except ValueError:
        return None

def _should_skip(label: str) -> bool:
    n = _norm(label)
    if not n or len(n) < 2:
        return True
    if n in SKIP_EXACT:
        return True
    if any(n.startswith(p) for p in SKIP_PREFIX):
        return True
    if any(n.endswith(s) for s in SKIP_SUFFIX):
        return True
    if re.match(r"^[\d\s\(\)\+\-\=\/]+$", n):
        return True
    if len(n.replace(" ", "")) <= 2:
        return True
    return False

def _row_type(label: str) -> str:
    n = _norm(label)
    if any(n.startswith(k) for k in TOTAL_KW) or "total general" in n:
        return "total"
    if any(n.startswith(k) for k in RESULT_KW):
        return "result"
    if any(n.startswith(k) for k in SECTION_KW):
        return "section"
    stripped = label.strip()
    if stripped and stripped == stripped.upper() and len(stripped) > 4:
        return "section"
    return "normal"

def _clean_label(label: str) -> str:
    label = label.strip()
    # Enlever lettre(s) isolée(s) en début (décorations verticales PDF)
    label = re.sub(r"^[A-Z]{1,2}\s+(?=[A-ZÀ-Ü])", "", label)
    label = re.sub(r"^(I{1,3}|IV|V|VI{1,3}|IX|X{1,3})\s+", "", label)
    return label.strip()


# ═══════════════════════════════════════════════════════════════════════════════
# EXTRACTION PAR COORDONNÉES BBOX
# ═══════════════════════════════════════════════════════════════════════════════

class WordExtractor:
    """
    Extrait les données d'une page PDF par coordonnées bbox.
    Ne dépend pas de extract_tables() — fonctionne avec tous les PDFs MCN.
    """

    def __init__(self, page, label_max_x_ratio: float = 0.40):
        self.page   = page
        self.pw     = page.width
        self.lmax   = self.pw * label_max_x_ratio
        self._words = page.extract_words(x_tolerance=3, y_tolerance=3)

    def _group_lines(self, y_tol: int = 3) -> dict:
        lines = defaultdict(list)
        for w in self._words:
            y_key = round(w['top'] / y_tol) * y_tol
            lines[y_key].append(w)
        return {y: sorted(ws, key=lambda w: w['x0']) for y, ws in lines.items()}

    def _col_tokens(self, ws: list, x_min: float, x_max: float) -> list:
        return [w['text'] for w in ws if x_min <= w['x0'] < x_max]

    def _col_bounds(self, mode: str) -> list:
        """
        Bornes (x_min, x_max) en proportion de la largeur de page.
        Calibrées sur PDFs MCN standard (A4 = 595pt).
        """
        pw = self.pw
        if mode == "actif":
            return [
                (pw * 0.440, pw * 0.560),   # Brut
                (pw * 0.560, pw * 0.680),   # Amortissements
                (pw * 0.680, pw * 0.790),   # Net exercice N
                (pw * 0.790, pw * 0.940),   # Net exercice N-1
            ]
        elif mode == "passif":
            return [
                (pw * 0.620, pw * 0.780),   # Exercice N
                (pw * 0.780, pw * 0.940),   # Exercice N-1
            ]
        elif mode == "cpc":
            return [
                (pw * 0.430, pw * 0.560),   # Propres à l'exercice
                (pw * 0.560, pw * 0.690),   # Exercices précédents
                (pw * 0.690, pw * 0.810),   # Totaux exercice N
                (pw * 0.810, pw * 0.940),   # Totaux exercice N-1
            ]
        return []

    def extract_rows(self, mode: str, skip_header_lines: int = 8) -> list:
        lines      = self._group_lines()
        col_bounds = self._col_bounds(mode)
        rows       = []
        seen       = set()

        for i, y in enumerate(sorted(lines.keys())):
            if i < skip_header_lines:
                continue
            ws = lines[y]
            label_tokens = [w['text'] for w in ws if w['x0'] < self.lmax]
            label = _clean_label(" ".join(label_tokens))

            if not label or _should_skip(label):
                continue
            key = _norm(label)
            if key in seen:
                continue
            seen.add(key)

            vals = [_parse_num_tokens(self._col_tokens(ws, xmin, xmax))
                    for xmin, xmax in col_bounds]

            row = {"label": label, "type": _row_type(label)}
            if mode == "actif":
                row.update({"brut": vals[0], "amort": vals[1],
                             "net_n": vals[2], "net_n1": vals[3]})
            elif mode == "passif":
                row.update({"val_n": vals[0], "val_n1": vals[1]})
            elif mode == "cpc":
                row.update({"propre_n": vals[0], "prec_n": vals[1],
                             "total_n": vals[2], "total_n1": vals[3]})
            rows.append(row)

        return rows


# ═══════════════════════════════════════════════════════════════════════════════
# PARSER PRINCIPAL
# ═══════════════════════════════════════════════════════════════════════════════

class PDFParser:

    def __init__(self, pdf_path: str):
        self.pdf   = pdfplumber.open(pdf_path)
        self.pages = self.pdf.pages
        self.n     = len(self.pages)

    def parse(self) -> dict:
        ranges = self._detect_ranges()
        return {
            "info":   self._extract_info(),
            "actif":  self._extract_section(ranges["actif"],  "actif"),
            "passif": self._extract_section(ranges["passif"], "passif"),
            "cpc":    self._extract_section(ranges["cpc"],    "cpc"),
            "pages":  self.n,
        }

    def close(self):
        try:
            self.pdf.close()
        except Exception:
            pass

    def _detect_ranges(self) -> dict:
        ranges = {"actif": [], "passif": [], "cpc": []}
        for i, page in enumerate(self.pages):
            t = _norm(page.extract_text() or "")
            if any(k in t for k in ACTIF_KW):
                ranges["actif"].append(i)
            if any(k in t for k in PASSIF_KW):
                ranges["passif"].append(i)
            if any(k in t for k in CPC_KW):
                ranges["cpc"].append(i)
        if not ranges["actif"]:
            ranges["actif"] = list(range(1, min(3, self.n)))
        if not ranges["passif"]:
            ranges["passif"] = list(range(2, min(5, self.n)))
        if not ranges["cpc"]:
            ranges["cpc"] = list(range(3, min(7, self.n)))
        return ranges

    def _extract_info(self) -> dict:
        info = {}
        for i in range(min(3, self.n)):
            text = self.pages[i].extract_text() or ""
            if not info.get("raison_sociale"):
                m = re.search(r"Raison sociale\s*:?\s*(.+)", text, re.IGNORECASE)
                if m:
                    info["raison_sociale"] = m.group(1).strip()
            if not info.get("identifiant_fiscal"):
                m = re.search(r"Identifiant [Ff]iscal\s*:?\s*(\d+)", text)
                if m:
                    info["identifiant_fiscal"] = m.group(1)
            if not info.get("taxe_professionnelle"):
                m = re.search(r"Taxe [Pp]rofessionnelle\s*:?\s*(\d+)", text)
                if m:
                    info["taxe_professionnelle"] = m.group(1)
            if not info.get("adresse"):
                m = re.search(r"Adresse\s*:?\s*(.+)", text, re.IGNORECASE)
                if m:
                    info["adresse"] = m.group(1).strip()
            if not info.get("exercice"):
                m = re.search(r"(\d{2}/\d{2}/\d{4})\s+au\s+(\d{2}/\d{2}/\d{4})", text)
                if m:
                    info["exercice"]       = f"Du {m.group(1)} au {m.group(2)}"
                    info["exercice_debut"] = m.group(1)
                    info["exercice_fin"]   = m.group(2)

        info.setdefault("raison_sociale", "")
        info.setdefault("identifiant_fiscal", "")
        info.setdefault("exercice", "")
        info.setdefault("exercice_fin", "")
        return info

    def _extract_section(self, page_indices: list, mode: str) -> list:
        all_rows, seen = [], set()
        for idx in page_indices:
            if idx >= self.n:
                continue
            for row in WordExtractor(self.pages[idx]).extract_rows(mode):
                key = _norm(row["label"])
                if key not in seen:
                    seen.add(key)
                    all_rows.append(row)
        return all_rows


# ═══════════════════════════════════════════════════════════════════════════════
# FORMATAGE EXCEL
# ═══════════════════════════════════════════════════════════════════════════════

def _border():
    s = Side(style='thin', color=C_BORDER)
    return Border(top=s, bottom=s, left=s, right=s)

def _fills(typ: str) -> tuple:
    if typ == "total":   return C_DARK_BLUE, C_WHITE, True
    if typ == "result":  return C_RESULT,    C_WHITE, True
    if typ == "section": return C_SECTION, C_DARK_BLUE, True
    if typ == "subtotal":return C_SUBTOTAL,C_DARK_BLUE, False
    return C_WHITE, "222222", False

def _cell(ws, row, col, value=None, bg=C_WHITE, fg="222222",
          bold=False, align="left", num_fmt=None,
          indent=0, size=9, wrap=True):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(name="Arial", size=size, bold=bold, color=fg)
    c.fill      = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal=align, vertical="center",
                            indent=indent, wrap_text=wrap)
    c.border    = _border()
    if num_fmt:
        c.number_format = num_fmt
    return c

def _header(ws, r, specs, height=28):
    ws.row_dimensions[r].height = height
    for cs, ce, text, bg in specs:
        if cs != ce:
            ws.merge_cells(start_row=r, start_column=cs,
                           end_row=r, end_column=ce)
        _cell(ws, r, cs, text, bg=bg, fg=C_WHITE, bold=True,
              align="center", size=10, wrap=True)

def _cover(ws, r, raison, exercice, if_num, title, ncols):
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
    _cell(ws, r, 1, title, bg=C_DARK_BLUE, fg=C_WHITE, bold=True,
          align="center", size=12)
    ws.row_dimensions[r].height = 26
    r += 1

    split = max(ncols - 2, 1)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=split)
    _cell(ws, r, 1, f"Raison sociale : {raison}",
          bg=C_LIGHT_BLUE, fg=C_DARK_BLUE, bold=True, size=9, indent=1)
    if ncols > 2:
        ws.merge_cells(start_row=r, start_column=split+1,
                       end_row=r, end_column=ncols)
        _cell(ws, r, split+1, f"IF : {if_num}",
              bg=C_LIGHT_BLUE, fg=C_DARK_BLUE, align="right", size=9)
    ws.row_dimensions[r].height = 16
    r += 1

    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
    _cell(ws, r, 1, f"Exercice : {exercice}",
          bg=C_GRAY_BG, fg="555555", size=9, indent=1)
    ws.row_dimensions[r].height = 14
    r += 1

    for c in range(1, ncols + 1):
        _cell(ws, r, c, bg=C_WHITE)
    ws.row_dimensions[r].height = 4
    return r + 1

def _data_row(ws, r, row_type, cells):
    bg, fg, bold = _fills(row_type)
    ws.row_dimensions[r].height = 15 if row_type == "normal" else 17
    for col, value, align, is_num in cells:
        _cell(ws, r, col, value, bg=bg, fg=fg, bold=bold, align=align,
              num_fmt=NUM_FMT if (is_num and value is not None) else None,
              indent=1 if (align == "left" and row_type == "normal") else 0)


# ── Feuilles ─────────────────────────────────────────────────────────────────

def _sheet_ident(wb, info):
    ws = wb.create_sheet("1 — Identification")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 50

    ws.merge_cells('A1:B1')
    _cell(ws, 1, 1, "PIÈCES ANNEXES À LA DÉCLARATION FISCALE",
          bg=C_DARK_BLUE, fg=C_WHITE, bold=True, align="center", size=13)
    ws.row_dimensions[1].height = 28

    ws.merge_cells('A2:B2')
    _cell(ws, 2, 1, "IMPÔTS SUR LES SOCIÉTÉS — Modèle Comptable Normal (loi 9-88)",
          bg=C_MED_BLUE, fg=C_WHITE, align="center", size=10)
    ws.row_dimensions[2].height = 18

    fields = [
        ("Raison sociale",       info.get("raison_sociale", "—")),
        ("Identifiant fiscal",   info.get("identifiant_fiscal", "—")),
        ("Taxe professionnelle", info.get("taxe_professionnelle", "—")),
        ("Adresse",              info.get("adresse", "—")),
        ("Exercice",             info.get("exercice", "—")),
    ]
    for i, (lbl, val) in enumerate(fields, 4):
        ws.row_dimensions[i].height = 18
        _cell(ws, i, 1, lbl, bg=C_LIGHT_BLUE, fg=C_DARK_BLUE,
              bold=True, size=9, indent=1)
        _cell(ws, i, 2, val, bg=C_WHITE, fg="222222", size=9, indent=1)


def _sheet_actif(wb, info, rows):
    ws = wb.create_sheet("2 — Bilan Actif")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 46
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18

    r = _cover(ws, 1, info.get("raison_sociale", ""),
               info.get("exercice", ""), info.get("identifiant_fiscal", ""),
               "BILAN ACTIF — Modèle Comptable Normal", 5)

    _header(ws, r, [
        (1, 1, "ACTIF",              C_DARK_BLUE),
        (2, 2, "BRUT",               C_MED_BLUE),
        (3, 3, "AMORT. & PROV.",     C_MED_BLUE),
        (4, 4, "NET — EXERCICE N",   C_DARK_BLUE),
        (5, 5, "NET — EXERCICE N-1", C_MED_BLUE),
    ])
    ws.freeze_panes = f"A{r+1}"
    r += 1

    for d in rows:
        _data_row(ws, r, d.get("type", "normal"), [
            (1, d["label"],      "left",  False),
            (2, d.get("brut"),   "right", True),
            (3, d.get("amort"),  "right", True),
            (4, d.get("net_n"),  "right", True),
            (5, d.get("net_n1"), "right", True),
        ])
        r += 1
    return len(rows)


def _sheet_passif(wb, info, rows):
    ws = wb.create_sheet("3 — Bilan Passif")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20

    r = _cover(ws, 1, info.get("raison_sociale", ""),
               info.get("exercice", ""), info.get("identifiant_fiscal", ""),
               "BILAN PASSIF — Modèle Comptable Normal", 3)

    _header(ws, r, [
        (1, 1, "PASSIF",       C_DARK_BLUE),
        (2, 2, "EXERCICE N",   C_DARK_BLUE),
        (3, 3, "EXERCICE N-1", C_MED_BLUE),
    ])
    ws.freeze_panes = f"A{r+1}"
    r += 1

    for d in rows:
        _data_row(ws, r, d.get("type", "normal"), [
            (1, d["label"],      "left",  False),
            (2, d.get("val_n"),  "right", True),
            (3, d.get("val_n1"), "right", True),
        ])
        r += 1

    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    c = ws.cell(row=r, column=1,
        value="(1) Capital personnel débiteur.  (2) Bénéficiaire (+) / Déficitaire (−).")
    c.font = Font(name="Arial", italic=True, size=8, color="888888")
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    return len(rows)


def _sheet_cpc(wb, info, rows):
    ws = wb.create_sheet("4 — CPC")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18

    r = _cover(ws, 1, info.get("raison_sociale", ""),
               info.get("exercice", ""), info.get("identifiant_fiscal", ""),
               "COMPTE DE PRODUITS ET CHARGES (Hors Taxes)", 5)

    _header(ws, r, [
        (1, 1, "DÉSIGNATION",           C_DARK_BLUE),
        (2, 2, "PROPRES À\nL'EXERCICE", C_MED_BLUE),
        (3, 3, "EXERCICES\nPRÉCÉDENTS", C_MED_BLUE),
        (4, 4, "TOTAUX\nEXERCICE N",    C_DARK_BLUE),
        (5, 5, "TOTAUX\nEXERCICE N-1",  C_MED_BLUE),
    ])
    ws.freeze_panes = f"A{r+1}"
    r += 1

    for d in rows:
        _data_row(ws, r, d.get("type", "normal"), [
            (1, d["label"],        "left",  False),
            (2, d.get("propre_n"), "right", True),
            (3, d.get("prec_n"),   "right", True),
            (4, d.get("total_n"),  "right", True),
            (5, d.get("total_n1"), "right", True),
        ])
        r += 1

    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    c = ws.cell(row=r, column=1,
        value="(1) Stock final − Stock initial : Augmentation (+) / Diminution (−).   "
              "(2) Achats revendus = Achats − Variation de stock.")
    c.font = Font(name="Arial", italic=True, size=8, color="888888")
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    return len(rows)


# ═══════════════════════════════════════════════════════════════════════════════
# POINT D'ENTRÉE PUBLIC
# ═══════════════════════════════════════════════════════════════════════════════

def convert(pdf_path: str, output_path: str) -> dict:
    """
    Extrait le PDF fiscal et génère un Excel structuré et formaté.
    Retourne : {info, tables, rows, pages}  — compatible app.py.
    """
    parser = PDFParser(pdf_path)
    try:
        data = parser.parse()
    finally:
        parser.close()

    info, actif, passif, cpc = (
        data["info"], data["actif"], data["passif"], data["cpc"]
    )

    wb = Workbook()
    wb.remove(wb.active)
    _sheet_ident(wb, info)
    n_a = _sheet_actif(wb, info, actif)
    n_p = _sheet_passif(wb, info, passif)
    n_c = _sheet_cpc(wb, info, cpc)
    wb.save(output_path)

    return {
        "info":   info,
        "tables": sum([bool(actif), bool(passif), bool(cpc)]),
        "rows":   n_a + n_p + n_c,
        "pages":  data["pages"],
    }
