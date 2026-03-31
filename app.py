"""
FiscalXL v2 — PDF Fiscal → Excel Pro
Nouvelle approche : injection dans template (formules déjà présentes)
"""

import streamlit as st
import tempfile, os
from pathlib import Path
import pandas as pd

from core.pdf_parser import PDFParser
from core.injector import TemplateInjector
from utils.validator import validate_pdf_structure_v2
from utils.logger import get_logger

logger = get_logger(__name__)

TEMPLATE_PATH = Path(__file__).parent / "template_fiscal.xlsx"

st.set_page_config(
    page_title="FiscalXL — PDF → Excel",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded",
)

st.markdown("""
<style>
.main-header {
    background: linear-gradient(135deg, #1F3864 0%, #2E75B6 100%);
    padding: 2rem 2.5rem; border-radius: 12px; margin-bottom: 1.5rem;
}
.main-header h1 { color: white; margin: 0; font-size: 2rem; }
.main-header p  { color: #BDD7EE; margin: 0.4rem 0 0; font-size: 0.95rem; }
.step-card {
    background: #f8f9fa; border-left: 4px solid #2E75B6;
    padding: 0.8rem 1rem; border-radius: 0 8px 8px 0; margin: 0.4rem 0;
}
.step-card strong { color: #1F3864; }
.step-card span   { color: #555; font-size: 0.88rem; }
.kpi-box {
    background: white; border: 1px solid #BDD7EE;
    border-radius: 8px; padding: 0.8rem; text-align: center;
}
.kpi-box .val { font-size: 1.3rem; font-weight: bold; color: #1F3864; }
.kpi-box .lbl { font-size: 0.75rem; color: #888; margin-top: 0.2rem; }
.success-box { background:#E2EFDA; border:1px solid #70AD47; border-radius:8px; padding:1rem 1.5rem; color:#375623; }
.warn-box    { background:#FFF2CC; border:1px solid #FFD700; border-radius:8px; padding:0.8rem 1.2rem; color:#7B5900; }
.error-box   { background:#FCE4D6; border:1px solid #C55A11; border-radius:8px; padding:1rem 1.5rem; color:#7B2C00; }
div[data-testid="stDownloadButton"] button {
    background: linear-gradient(135deg, #1F3864, #2E75B6);
    color:white; border:none; padding:0.8rem 2.5rem;
    font-size:1.05rem; border-radius:8px; width:100%;
}
</style>
""", unsafe_allow_html=True)

st.markdown("""
<div class="main-header">
    <h1>📊 FiscalXL</h1>
    <p>Convertisseur PDF → Excel · Pièces annexes IS (Modèle Comptable Normal, loi 9-88)</p>
</div>
""", unsafe_allow_html=True)

with st.sidebar:
    st.markdown("### ⚙️ Options")
    show_preview = st.toggle("Aperçu des données", value=True)
    show_debug   = st.toggle("Mode débogage",      value=False)
    st.markdown("---")
    st.markdown("### 📋 Format PDF accepté")
    st.markdown(
        "Le PDF doit contenir :\n"
        "- **Bilan Actif** (immobilisé + circulant)\n"
        "- **Bilan Passif** (capitaux propres + dettes)\n"
        "- **CPC** (Compte de Produits et Charges)\n\n"
        "Les pages sont détectées **automatiquement**."
    )
    st.markdown("---")
    if not TEMPLATE_PATH.exists():
        st.error("⚠️ Template introuvable")
    else:
        st.success("✅ Template chargé")
    st.caption("FiscalXL v2 · MCN loi 9-88")

col_up, col_steps = st.columns([3, 2])
with col_up:
    st.markdown("### 📂 Importer le PDF")
    uploaded = st.file_uploader("Glissez-déposez ou cliquez", type=["pdf"])

with col_steps:
    st.markdown("### 🔄 Pipeline")
    for step, desc in [
        ("1 · Lecture PDF",         "pdfplumber extrait tableaux et texte"),
        ("2 · Détection sections",  "Pages Actif / Passif / CPC détectées auto"),
        ("3 · Injection template",  "Valeurs → cellules exactes du modèle"),
        ("4 · Excel avec formules", "Totaux et résultats calculés auto"),
    ]:
        st.markdown(f'<div class="step-card"><strong>{step}</strong><br><span>{desc}</span></div>', unsafe_allow_html=True)

if uploaded:
    st.markdown("---")

    if not TEMPLATE_PATH.exists():
        st.markdown('<div class="error-box">❌ <strong>Template manquant.</strong></div>', unsafe_allow_html=True)
        st.stop()

    with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp:
        tmp.write(uploaded.getbuffer())
        pdf_path = tmp.name
    output_path = pdf_path.replace(".pdf", "_out.xlsx")
    excel_bytes = None  # sera rempli avant le finally

    try:
        progress = st.progress(0)
        status   = st.empty()

        # Étape 1 — Validation
        status.info("🔍 Étape 1/4 — Validation du PDF...")
        progress.progress(10)
        parser = PDFParser(pdf_path)
        validation = validate_pdf_structure_v2(parser)
        if not validation["valid"]:
            st.markdown(f'<div class="error-box">⚠️ {validation["message"]}</div>', unsafe_allow_html=True)
            st.stop()

        meta = validation["meta"]
        c1, c2, c3, c4 = st.columns(4)
        for col, (lbl, val) in zip([c1,c2,c3,c4], [
            ("Raison Sociale",    (meta.get("raison_sociale") or "—")[:22]),
            ("Identifiant Fiscal", meta.get("identifiant_fiscal") or "—"),
            ("Fin exercice",      meta.get("exercice_fin") or "—"),
            ("Pages",             str(meta.get("pages","—"))),
        ]):
            col.markdown(f'<div class="kpi-box"><div class="val">{val}</div><div class="lbl">{lbl}</div></div>', unsafe_allow_html=True)

        st.markdown("<br>", unsafe_allow_html=True)
        progress.progress(20)

        # Étape 2 — Extraction
        status.info("📄 Étape 2/4 — Extraction des valeurs PDF...")
        progress.progress(40)
        with st.spinner("Parsing..."):
            extracted = parser.parse()
        progress.progress(60)

        # Afficher les pages détectées
        if show_debug:
            with st.expander("🗂️ Pages détectées automatiquement"):
                st.json(parser._ranges)
            with st.expander("🐛 Données brutes extraites"):
                st.json({k: {str(lbl): v for lbl,v in d.items()} if isinstance(d, dict) else d
                         for k,d in extracted.items()})

        # Étape 3 — Injection
        status.info("🔗 Étape 3/4 — Injection dans le template...")
        progress.progress(70)
        with st.spinner("Injection..."):
            injector = TemplateInjector(str(TEMPLATE_PATH))
            stats = injector.inject(extracted, output_path)
        progress.progress(88)

        # Étape 4 — Vérification + lecture en mémoire
        status.info("✅ Étape 4/4 — Vérification...")
        import openpyxl
        wb_check = openpyxl.load_workbook(output_path)
        n_formulas = sum(
            1 for ws in wb_check.worksheets
            for row in ws.iter_rows()
            for c in row
            if isinstance(c.value, str) and c.value.startswith("=")
        )
        progress.progress(100)

        # ── LECTURE EN MÉMOIRE avant suppression ──
        with open(output_path, "rb") as f:
            excel_bytes = f.read()

        status.empty()

        st.markdown(f"""
        <div class="success-box">
            ✅ <strong>Fichier Excel généré !</strong>
            &nbsp;·&nbsp; {len(wb_check.sheetnames)} feuilles
            &nbsp;·&nbsp; {n_formulas} formules intactes
            &nbsp;·&nbsp; {stats['injected']} valeurs injectées
        </div>""", unsafe_allow_html=True)

        if stats.get("not_found"):
            nf = stats["not_found"]
            st.markdown(f'<div class="warn-box">⚠️ <strong>{len(nf)} postes non mappés</strong> : {" · ".join(nf[:6])}</div>', unsafe_allow_html=True)

        st.markdown("<br>", unsafe_allow_html=True)

        if show_preview:
            with st.expander("👁️ Aperçu des valeurs extraites", expanded=True):
                tabs = st.tabs(["ℹ️ Infos", "📋 Bilan Actif", "📋 Bilan Passif", "📈 CPC"])

                with tabs[0]:
                    for k, v in extracted.get("info", {}).items():
                        if k != "pages":
                            st.markdown(f"**{k.replace('_',' ').title()}** : {v}")

                with tabs[1]:
                    av = extracted.get("actif_values", {})
                    if av:
                        df = pd.DataFrame(
                            [(k, *(f"{x:.2f}" if x is not None else "—" for x in (v+[None,None,None])[:3]))
                             for k, v in av.items()],
                            columns=["Poste","Brut","Amort.","Net N-1"])
                        st.dataframe(df, use_container_width=True, height=300)

                with tabs[2]:
                    pv = extracted.get("passif_values", {})
                    if pv:
                        df = pd.DataFrame(
                            [(k, *(f"{x:.2f}" if x is not None else "—" for x in (v+[None,None])[:2]))
                             for k, v in pv.items()],
                            columns=["Poste","Exercice N","Exercice N-1"])
                        st.dataframe(df, use_container_width=True, height=300)

                with tabs[3]:
                    cv = extracted.get("cpc_values", {})
                    if cv:
                        df = pd.DataFrame(
                            [(k, *(f"{x:.2f}" if x is not None else "—" for x in (v+[None,None,None])[:3]))
                             for k, v in cv.items()],
                            columns=["Désignation","Propre N","Exerc. Préc.","Total N-1"])
                        st.dataframe(df, use_container_width=True, height=300)

        st.markdown("### ⬇️ Télécharger")
        fname = Path(uploaded.name).stem + "_fiscal.xlsx"
        st.download_button(
            "📥 Télécharger le fichier Excel",
            data=excel_bytes,   # bytes en mémoire — pas de file handle
            file_name=fname,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    except Exception as e:
        logger.exception("Erreur pipeline")
        st.markdown(f'<div class="error-box">❌ <strong>Erreur :</strong> <code>{e}</code></div>', unsafe_allow_html=True)
        if show_debug:
            import traceback; st.code(traceback.format_exc())
    finally:
        # Suppression sécurisée des fichiers temporaires
        for f in [pdf_path, output_path]:
            try:
                if os.path.exists(f):
                    os.unlink(f)
            except Exception:
                pass

else:
    st.markdown("""
    <div style="text-align:center;padding:3rem;color:#888;
        border:2px dashed #BDD7EE;border-radius:12px;background:#f8fafd;margin-top:1rem;">
        <div style="font-size:3rem;">📄</div>
        <h3 style="color:#2E75B6;">Importez un PDF pour commencer</h3>
        <p>Pièces annexes à la déclaration IS — Modèle Comptable Normal · Détection automatique des sections</p>
    </div>""", unsafe_allow_html=True)
